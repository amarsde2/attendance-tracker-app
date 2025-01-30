"""

THIS IS ATTENDANCE TRACKER APPLICATION IN WHICH YOU CAN SETUP YOUR SUBJECTS FOR WHICH YOU WANT TO TRACK ATTENDANCE 
YOU CAN ALSO ADD YOUR CLASS STUDENTS THEIR INFO LIKE ROLL NUMBER AND EMAIL SO YOU CAN NOTIFY THEM ABOUT THEIR ABSENT
WE CAN ALSO NOTIFY FACULITY MEMBERS ABOUT THEIR CLASS ATTENDANCE 

AUTHOR: ER. AMAR KUMAR
PROFESSIONAL: SENIOR SOFTWARE ENGINNER 

"""

import openpyxl
from openpyxl import Workbook
import mailer
import sheet as asheet

class AttendanceTracker:

    # PROPERTY TO MANGE SUBJECTS AND USER CHOICE AND SHOULD PROCESS MORE ENTRIES
    _shouldProcess = True
    _validSubjects = [] 
    _choice = 0
   
    def __init__(self):
        '''
        SET UP XLSX SHEET FOR STORE ATTENDANCE DETAILS
        '''
        try:
            self._db = openpyxl.load_workbook("attendance.xlsx")
            self._sheet = self._db.active
    
        except FileNotFoundError:
            # SETUP NEW SHEET IF NOT ALREAY EXIST

            self._db = Workbook()
            self._sheet = self._db.active
            self._sheet.title = "Sheet1"
            self._db.save("attendance.xlsx")
            asheet.setupSheet(self._sheet, self._db)

        except Exception as e:
            self._db = None
            self._sheet = None
            return
                


    def _choiceListPlay(self):
        """
        FUNCTION TO REMEMBER USER CHOICE TO PERFORM OPERATIONS
        """

        print("Type 1 For add a new student in sheet: ")
        print("TYpe 2 to add absent students: ")
        print("Type 3 to reset sheet attendance: ")
        print("Type 4 to exit: ")

        self._choice = int(input("Enter your choice: "))

        while self._choice not in [1,2,3,4]:
            print("Enter valid option ")
            self._choice = int(input("Please Enter Valid Choice: "))

    
     
    def _fetchSubjects(self):
        """
        FUNCTION TO FETCH VALID SUBJECTS FROM SHEET
        """
        for idx, subject in enumerate(self._sheet[1], start=1):
            if idx > 2:
               self._validSubjects.append(subject.value)


    def _inputMessageInstruction(self):

        """
        FUNCTION TO GIVE HINT TO USERS FOR PICK SUBJECT 
        """
        print("Please following instruction to enter entries for attendance")         
        for index, i in enumerate(self._validSubjects,start=1):
            print(f"Type {index} to choose {i} as subject")


   
    def _insertAbsentEntries(self, subject, absents, rollNumber):
    
        """
        FUNCTION TO ADD ABSENTS FOR STUDENTS
        """
        for i in range(absents):
            for r in range(1, self._sheet.max_row):
    
                if  self._sheet.cell(row=r, column=1).value == str(rollNumber[i]): 
                        asheet.saveAttendace(self._sheet, self._db, r,  subject + 2)
                        
              
    def process(self):
        """
        MAIN ENTRY FUNCTION TO HANDLE APP FUNCTIONALITY
        """
        if not self._db or not self._sheet:
            print("Cannot process. Workbook or sheet is not initialized.")
            return
    
        self._choiceListPlay()
        self._fetchSubjects()
        
        if self._choice == 1:
            asheet.addStudent(self._sheet, self._db, self._validSubjects)
        
        elif self._choice == 2:
            self._inputMessageInstruction()
            try:
                while self._shouldProcess:

                    subj = int(input("Enter subject number for which you want to enter absent: "))
                    while subj < 1 or subj > len(self._validSubjects):
                        print("Please enter valid number: ")
                        subj = input("Enter subject number for which you want to enter absent: ")

                    no_of_absent = int(input("Enter a number of people absent: "))
                    roll_number = []
            
                    for _ in range(0, no_of_absent):
                        roll_number.append(int(input("Enter roll number of student: ")))
                    
                    self._insertAbsentEntries(subj, no_of_absent, roll_number)
                    self._shouldProcess = int(input("If you want more record type 1 or 0 for exit "))

                #mailer.notifyStaffMembers(self._sheet, self._validSubjects)

            except KeyboardInterrupt:
                print("You can use it whenever you want your entries are saved for future reference..")
                exit(0)
        elif self._choice == 3:
            asheet.resetAttendance(self._sheet)
        else:
            print("Bye, See you again...")
            exit(0)

            
if __name__ == "__main__":
    app = AttendanceTracker()
    app.process()
